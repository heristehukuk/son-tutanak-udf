
import re, io
from uuid import uuid4
from pathlib import Path
from openpyxl import load_workbook
from app.database_layer import repos
from app.auth.service import now
from app.documents.engine import turkce_sayi_yazi

TEMPLATE_PATH = Path(__file__).resolve().parents[1] / "templates" / "xlsx" / "harcama_pusulasi.xlsx"

# Dosya Türü (serbest metin) -> sabit kategori anahtarı. Anahtar kelime eşleştirmesiyle çalışır;
# eşleşme bulunamazsa 'diger' kategorisine düşer (Madde 8: tarifede yazılı olmayan haller).
CATEGORY_KEYWORDS = [
    ("kira", ["kira", "tahliye"]),
    ("ticari", ["ticari", "ticaret"]),
    ("ortakligin_giderilmesi", ["ortaklığın giderilmesi", "ortaklik"]),
    ("is", ["iş", "işçi", "işveren", "kıdem", "ihbar"]),
    ("tuketici", ["tüketici"]),
]
CATEGORY_LABELS = {
    "kira": "Kira", "ticari": "Ticari", "ortakligin_giderilmesi": "Ortaklığın Giderilmesi",
    "is": "İşçi-İşveren", "tuketici": "Tüketici", "diger": "Diğer",
}

def detect_category(dosya_turu_text):
    """Serbest metin 'Dosya Türü' alanından sabit tarife kategorisini tespit eder."""
    # NOT: Python'un .lower()'ı Türkçe büyük 'İ' harfini yanlış çevirir (birleşik nokta karakteri üretir).
    # Önce Türkçe kurala göre elle çeviriyoruz, sonra normal .lower() uyguluyoruz.
    t = (dosya_turu_text or "").strip().replace("İ", "i").replace("I", "ı").lower()
    if not t: return "diger"
    for key, words in CATEGORY_KEYWORDS:
        if any(w in t for w in words):
            return key
    return "diger"

def lookup_unit_price(category, taraf_sayisi, year=None):
    """fee_tariffs tablosundan kategori + taraf sayısına uyan birim fiyatı bulur.
    Bulunamazsa None döner (admin panelinden tarife eklenmesi gerekir)."""
    found = repos.tariffs.find_matching(category, taraf_sayisi, year)
    return found["unit_price"] if found else None

def _tr_title(s):
    """Python'un .title()'ı Türkçe İ/I kuralını bilmez (örn. 'iki' -> 'Iki' yanlış, 'İki' doğru).
    Her kelimenin ilk harfini Türkçe kurala göre büyütür."""
    def cap(word):
        if not word: return word
        first = "İ" if word[0] == "i" else word[0].upper()
        return first + word[1:]
    return " ".join(cap(w) for w in s.split(" "))

def turkce_para_yazi(amount):
    """Örn: 4680 -> 'Dört Bin Altı Yüz Seksen Türk Lirası' (Title Case, kuruş varsa 'X Kuruş' eklenir)."""
    tl = int(amount)
    kurus = round((amount - tl) * 100)
    words = _tr_title(turkce_sayi_yazi(tl)) if tl else "Sıfır"
    result = f"{words} Türk Lirası"
    if kurus:
        result += f" {_tr_title(turkce_sayi_yazi(kurus))} Kuruş"
    return result

def build_harcama_pusulasi(*, daire, dosya_turu_text, basvuru_no, taraf_sayisi, arabulucu_adi, arabulucu_tc, arabulucu_iban, year=None):
    """Harcama Pusulası xlsx'ini şablon üzerinden doldurup bytes olarak döner.
    Dönüş: (xlsx_bytes, uyari) -- uyari, tarife bulunamadıysa açıklayıcı bir mesajdır (None ise sorun yok)."""
    category = detect_category(dosya_turu_text)
    unit_price = lookup_unit_price(category, taraf_sayisi, year)
    uyari = None
    if unit_price is None:
        unit_price = 0.0
        uyari = (f"'{CATEGORY_LABELS.get(category,category)}' kategorisi ve {taraf_sayisi} taraf için tarifede "
                 f"kayıtlı bir birim fiyat bulunamadı. Birim Fiyatı/Tutarı 0 olarak bırakıldı; lütfen admin "
                 f"panelinden tarifeyi ekleyip tekrar üretin.")

    wb = load_workbook(TEMPLATE_PATH, data_only=False)
    ws = wb["Sayfa1"]

    label = CATEGORY_LABELS.get(category, category)
    ws["A3"] = f"Dairesi : {daire or ''}"
    ws["E6"] = f"1 Adet ({label} Dava Şartı Uyuşmazlığı - {taraf_sayisi} taraflı)"
    ws["E7"] = unit_price
    ws["E8"] = unit_price
    ws["A9"] = turkce_para_yazi(unit_price)
    ws["B10"] = f"Ankara Arabuluculuk Bürosunun {basvuru_no or ''} numaralı dosyasına istinaden arabuluculuk ücreti"
    ws["D16"] = arabulucu_tc or ""
    ws["D17"] = f"Arabulucu {arabulucu_adi or ''}".strip()
    ws["A18"] = "İban"
    ws["C18"] = ":"
    ws["D18"] = arabulucu_iban or ""
    ws["D18"].number_format = "@"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), uyari

# --- Admin: tarife tablosu yönetimi ---
def list_tariffs():
    return repos.tariffs.list_all()

def add_tariff(category, min_parties, max_parties, unit_price, year):
    repos.tariffs.create({
        "category":category,"category_label":CATEGORY_LABELS.get(category,category),
        "min_parties":min_parties,"max_parties":max_parties,"unit_price":unit_price,
        "year":year,"updated_at":now().isoformat(),
    })

def delete_tariff(tariff_id):
    repos.tariffs.delete(tariff_id)

def seed_known_tariffs():
    """Şu ana kadar doğrulanmış örnek rakamları (kullanıcının verdiği örnek pusulalardan) tohumlar.
    Tabloda hiç satır yoksa çalışır; tam resmi tarife Excel'i yüklenince admin panelinden tamamlanmalı."""
    if repos.tariffs.list_all(): return
    add_tariff("kira", 2, 2, 4680, 2026)
    add_tariff("ticari", 3, 3, 6400, 2026)
